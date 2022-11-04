from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Maternity Data")
root.geometry('800x700+300+20')

#resizable() method is used to allow Tkinter root window to change it’s size
# according to the users need as well we can prohibit resizing of the Tkinter window.
#So, basically, if user wants to create a fixed size window, this method can be used.
root.resizable(False,False)
root.configure(bg="#326273")

file=pathlib.Path('Backend_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Submission Date"
    sheet['B1']='Admissions'
    sheet['C1']='Deliveries'
    sheet['D1']="Male"
    sheet['E1']='Female'
    sheet['F1']='Below_14yrs'
    sheet['G1']='Above_19yrs'
    sheet['H1']='PPH'
    sheet['I1']='Neonatal_Deaths'
    sheet['J1']='Month'
    sheet['k1']='Year'


    
    file.save('Backnd_data.xlsx')

#def check():
#if name=nameValue.get():
 #   pass
#else:
 #   messagebox.showinfo('info','detail added')
  #  nameValue.set('')
    #name=nameValue.get()
    #contact=contactValue.get()
    #age=AgeValue.get()
    #gender=gender_combobox.get()
    #address=addressEntry.get(1.0,END)


def submit():
    Submission_Date=submissiondateValue.get()
    Admissions=admissionsValue.get()
    Deliveries=deliveriesValue.get()
    Male=maleValue.get()
    Female=femaleValue.get()
    Below_14yrs=lessfourteenValue.get()
    Above_19yrs=greaternineteenValue.get()
    PPH=pphValue.get()
    Neonatal_Deaths=neonatalValue.get()
    Year=year_combobox.get()
    Month=month_combobox.get()
    #address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('Backnd_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=Submission_Date)
    sheet.cell(column=2,row=sheet.max_row,value=Admissions)
    sheet.cell(column=3,row=sheet.max_row,value=Deliveries)
    sheet.cell(column=4,row=sheet.max_row,value=Male)
    sheet.cell(column=5,row=sheet.max_row,value=Female)
    sheet.cell(column=6,row=sheet.max_row,value=Below_14yrs)
    sheet.cell(column=7,row=sheet.max_row,value=Above_19yrs)
    sheet.cell(column=8,row=sheet.max_row,value=PPH)
    sheet.cell(column=9,row=sheet.max_row,value=Neonatal_Deaths)
    sheet.cell(column=10,row=sheet.max_row,value=Year)
    sheet.cell(column=11,row=sheet.max_row,value=Month)


    file.save(r'Backnd_data.xlsx')

    #messagebox.showinfo('info','detail added')
    #nameValue.set('')
    #contactValue.set('')
    #AgeValue.set('')
    #addressEntry.delete(1.0,END)

    yearValue.set('')
    monthValue.set('')
    submissiondateValue.set('')
    admissionsValue.set('')
    deliveriesValue.set('')
    maleValue.set('')
    femaleValue.set('')
    lessfourteenValue.set('')
    greaternineteenValue.set('')
    pphValue.set('')
    neonatalValue.set('')




def Clear():
    #nameValue.set('')
    #contactValue.set('')
    #AgeValue.set('')
    #addressEntry.delete(1.0,END)
    yearValue.set('')
    monthValue.set('')
    submissiondateValue.set('')
    admissionsValue.set('')
    deliveriesValue.set('')
    maleValue.set('')
    femaleValue.set('')
    lessfourteenValue.set('')
    greaternineteenValue.set('')
    pphValue.set('')
    neonatalValue.set('')




#icon
icon_image=PhotoImage(file='cfk.png')
root.iconphoto(False,icon_image)

#heading
Label(root,text=("please fill out this Entry form"),font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#label
Label(root,text="Year",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="Month",font=23,bg="#326273",fg="#fff").place(x=400,y=100)
Label(root,text="Submission_Date",font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text="Admissions",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="Deliveries",font=23,bg="#326273",fg="#fff").place(x=50,y=250)
Label(root,text="Male",font=23,bg="#326273",fg="#fff").place(x=50,y=300)
Label(root,text="Female",font=23,bg="#326273",fg="#fff").place(x=50,y=350)
Label(root,text="Below_14yrs",font=23,bg="#326273",fg="#fff").place(x=50,y=400)
Label(root,text="Above_19yrs",font=23,bg="#326273",fg="#fff").place(x=50,y=450)
Label(root,text="PPH",font=23,bg="#326273",fg="#fff").place(x=50,y=500)
Label(root,text="Neonatal_Deaths",font=23,bg="#326273",fg="#fff").place(x=50,y=550)


#Entry
yearValue=StringVar()
monthValue=StringVar()
submissiondateValue=StringVar()
admissionsValue=StringVar()
deliveriesValue=StringVar()
maleValue=StringVar()
femaleValue=StringVar()
lessfourteenValue=StringVar()
greaternineteenValue=StringVar()
pphValue=StringVar()
neonatalValue=StringVar()

submissiondateEntry=Entry(root,textvariable=submissiondateValue,width=45,bd=2,font=20)
admissionsEntry=Entry(root,textvariable=admissionsValue,width=45,bd=2,font=20)
deliveriesEntry=Entry(root,textvariable=deliveriesValue,width=45,bd=2,font=20)
maleEntry=Entry(root,textvariable=maleValue,width=45,bd=2,font=20)
femaleEntry=Entry(root,textvariable=femaleValue,width=45,bd=2,font=20)
lessfourteenEntry=Entry(root,textvariable=lessfourteenValue,width=45,bd=2,font=20)
greaternineteenEntry=Entry(root,textvariable=greaternineteenValue,width=45,bd=2,font=20)
pphEntry=Entry(root,textvariable=pphValue,width=45,bd=2,font=20)
neonatalEntry=Entry(root,textvariable=neonatalValue,width=45,bd=2,font=20)

#month dropdown
month_combobox=Combobox(root,values=['January','February','March','April','May','June','July','August','September','October','November','December'],font='arial 14',state='r',width=14)
month_combobox.place(x=500,y=100)
month_combobox.set('January')

#month dropdown
year_combobox=Combobox(root,values=['2023','2022','2021','2020','2019','2018','2017','2016','2015'],font='arial 14',state='r',width=14)
year_combobox.place(x=200,y=100)
year_combobox.set('2022')

#addressEntry=Text(root,width=50,height=4,bd=2)

submissiondateEntry.place(x=200,y=150)
admissionsEntry.place(x=200,y=200)
deliveriesEntry.place(x=200,y=250)
maleEntry.place(x=200,y=300)
femaleEntry.place(x=200,y=350)
lessfourteenEntry.place(x=200,y=400)
greaternineteenEntry.place(x=200,y=450)
pphEntry.place(x=200,y=500)
neonatalEntry.place(x=200,y=550)


Button(root,text='Submit',bg='#326273',fg='white',width=15,height=2,command=submit).place(x=200,y=600)
Button(root,text='Clear',bg='#326273',fg='white',width=15,height=2,command=Clear).place(x=340,y=600)
Button(root,text='Exit',bg='#326273',fg='white',width=15,height=2,command=lambda:root.destroy()).place(x=480,y=600)


#Root. mainloop() is simply a method in the main window that executes what we wish to execute in an application
# (lets Tkinter to start running the application). 
root.mainloop()